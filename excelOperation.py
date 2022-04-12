
import openpyxl
filename = 'C:\\Users\\selfg\\PycharmProjects\\selleniumPythonProject\\IDpass.xlsx'
# Load/ open excel file
workbook = openpyxl.load_workbook(filename)

# Locate/get perticular sheet
sheet = workbook["Sheet1"]                  # workbook.get_sheet_by_name("Sheet1")

# Read data from sheet
print(sheet.cell(row=1, column=1).value)
print(sheet.cell(row=3, column=2).value)

temp = sheet.cell(row=1, column=1).value

sheet.cell(row=7, column=1).value = "Vaishnavi"
sheet.cell(row=7, column=2).value = "VaishnaviPass"

print(sheet.cell(row=7, column=1).value)
print(sheet.cell(row=7, column=2).value)

workbook.save(filename)



